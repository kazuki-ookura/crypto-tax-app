#!/usr/bin/env python3
"""
国税庁「暗号資産の計算書（総平均法用）」自動記入スクリプト
002.xlsx に指定年の取引データを記入する
使い方: python fill_nta_excel.py [--year 2025]
"""
import os
import sys
import csv
import copy
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# --year: 指定なければ前年
if "--year" in sys.argv:
    idx = sys.argv.index("--year")
    YEAR = int(sys.argv[idx + 1])
else:
    YEAR = datetime.now().year - 1

SRC = os.path.join(SCRIPT_DIR, "data", "NTA", "002.xlsx")
DST = os.path.join(SCRIPT_DIR, f"{YEAR}_暗号資産の計算書.xlsx")
TX_CSV = os.path.join(SCRIPT_DIR, f"{YEAR}_crypto_transactions.csv")

print(f"対象年度: {YEAR}年")

# --- 1. Load and aggregate transaction data ---
with open(TX_CSV, "r", encoding="utf-8-sig") as f:
    rows = list(csv.DictReader(f))

# Section2: exchange buy/sell per currency
sec2 = defaultdict(lambda: defaultdict(lambda: {"bq": Decimal(0), "ba": Decimal(0), "sq": Decimal(0), "sa": Decimal(0)}))
# Section3: other transactions per currency
sec3 = defaultdict(list)

for row in rows:
    cat, cur, ex, sub = row["取引分類"], row["通貨"], row["取引所"], row["取引内容"]
    if cur == "JPY":
        continue
    try:
        qty = Decimal(row["数量"]) if row["数量"] else Decimal(0)
    except:
        qty = Decimal(0)
    try:
        amt = Decimal(row["金額(円)"]) if row["金額(円)"] else Decimal(0)
    except:
        amt = Decimal(0)

    ds = row["日付"]
    mo = ds[5:7].lstrip("0")
    da = ds[8:10].lstrip("0")

    if cat == "売買":
        if "買" in sub:
            sec2[cur][ex]["bq"] += qty
            sec2[cur][ex]["ba"] += abs(amt)
        elif "売" in sub:
            sec2[cur][ex]["sq"] += qty
            sec2[cur][ex]["sa"] += abs(amt)
    elif cat in ("レンディング報酬", "ステーキング報酬", "紹介報酬", "その他報酬", "受取"):
        sec3[cur].append({
            "mo": mo, "da": da, "partner": ex, "desc": sub,
            "bq": qty, "ba": abs(amt), "sq": Decimal(0), "sa": Decimal(0)
        })
    elif cat == "手数料返金":
        sec3[cur].append({
            "mo": mo, "da": da, "partner": ex, "desc": "手数料返金",
            "bq": Decimal(0), "ba": abs(amt), "sq": Decimal(0), "sa": Decimal(0)
        })

# Also include GMO staking deposits and campaign rewards from the raw data
gmo_path = os.path.join(SCRIPT_DIR, "data", "GMOCoin", f"{YEAR}_trading_report.csv")
with open(gmo_path, "r", encoding="utf-8-sig") as f:
    gmo_rows = list(csv.DictReader(f))

from datetime import datetime
for row in gmo_rows:
    ds = row.get("日時", "").strip()
    if not ds:
        continue
    dt = datetime.strptime(ds, "%Y/%m/%d %H:%M")
    if dt.year != YEAR:
        continue
    seisan = row.get("精算区分", "").strip()
    if seisan != "暗号資産預入・送付":
        continue
    direction = row.get("授受区分", "").strip()
    if direction != "預入":
        continue
    memo = row.get("送付先/送付元", "").strip()
    cur = row.get("銘柄名", "").strip()
    qty = Decimal(row.get("数量", "0").replace(",", ""))
    if memo in ("ステーキング", "キャンペーン") and qty > 0:
        sec3[cur].append({
            "mo": str(dt.month), "da": str(dt.day),
            "partner": "GMOコイン", "desc": memo,
            "bq": qty, "ba": Decimal(0),
            "sq": Decimal(0), "sa": Decimal(0)
        })

# Define which currencies to fill and their sheet order
currencies = ["BTC", "ETH", "XRP", "DOGE", "BAT", "ATOM", "TRX", "SOL", "BNB"]
sheet_names = ["計算書①", "計算書②", "計算書③", "計算書④", "計算書⑤", "計算書⑥", "計算書⑦", "計算書⑧", "計算書⑨"]

# --- 2. Load workbook and fill ---
wb = load_workbook(SRC)

for i, cur in enumerate(currencies):
    ws = wb[sheet_names[i]]
    print(f"\n=== {sheet_names[i]}: {cur} ===")

    # Header: crypto name in U6
    ws["U6"] = cur

    # Section 2: Exchange buy/sell (rows 12-16)
    exchanges = list(sec2[cur].items())
    for j, (ex_name, d) in enumerate(exchanges[:5]):
        r = 12 + j
        ws.cell(row=r, column=3, value=ex_name)  # C: exchange name
        if d["bq"]:
            ws.cell(row=r, column=21, value=float(d["bq"]))  # U: buy qty
        if d["ba"]:
            ws.cell(row=r, column=30, value=int(d["ba"]))     # AD: buy amount
        if d["sq"]:
            ws.cell(row=r, column=39, value=float(d["sq"]))   # AM: sell qty
        if d["sa"]:
            ws.cell(row=r, column=48, value=int(d["sa"]))     # AV: sell amount
        print(f"  Section2 row{r}: {ex_name} buy={d['bq']}({d['ba']}円) sell={d['sq']}({d['sa']}円)")

    # Section 3: Other transactions (rows 23-35, up to 13 entries)
    other_txs = sec3.get(cur, [])
    # Sort by month/day
    other_txs.sort(key=lambda t: (int(t["mo"]), int(t["da"])))

    for j, t in enumerate(other_txs[:13]):
        r = 23 + j
        ws.cell(row=r, column=3, value=int(t["mo"]))   # C: month
        ws.cell(row=r, column=5, value=int(t["da"]))    # E: day
        ws.cell(row=r, column=7, value=t["partner"])     # G: partner
        ws.cell(row=r, column=13, value=t["desc"])       # M: description
        if t["bq"]:
            ws.cell(row=r, column=21, value=float(t["bq"]))  # U: buy qty
        if t["ba"]:
            ws.cell(row=r, column=30, value=int(t["ba"]))     # AD: buy amount
        if t["sq"]:
            ws.cell(row=r, column=39, value=float(t["sq"]))   # AM: sell qty
        if t["sa"]:
            ws.cell(row=r, column=48, value=int(t["sa"]))     # AV: sell amount
        print(f"  Section3 row{r}: {t['mo']}/{t['da']} {t['partner']} {t['desc']} buy={t['bq']}({t['ba']}円)")

    # Section 4: Year-start balance (N41, N42) = 0 for now
    # User must fill in if they had pre-YEAR holdings
    ws.cell(row=41, column=14, value=0)  # N41: year-start qty
    ws.cell(row=42, column=14, value=0)  # N42: year-start amount
    print(f"  Section4: 年始残高=0（要確認）")

# --- 3. Create summary sheet at the front ---
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

summary = wb.create_sheet("サマリー", 0)  # Insert at position 0

# Column widths
summary.column_dimensions["A"].width = 14
summary.column_dimensions["B"].width = 22
summary.column_dimensions["C"].width = 18
summary.column_dimensions["D"].width = 18
summary.column_dimensions["E"].width = 18

title_font = Font(name="Meiryo", size=14, bold=True)
header_font = Font(name="Meiryo", size=10, bold=True, color="FFFFFF")
cell_font = Font(name="Meiryo", size=10)
num_fmt = '#,##0'
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
sub_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))

# Title
summary.merge_cells("A1:E1")
c = summary["A1"]
c.value = f"{YEAR}年 暗号資産 所得サマリー"
c.font = title_font
c.alignment = Alignment(horizontal="center")

# --- Section A: 売買損益（総平均法）---
row = 3
summary.cell(row=row, column=1, value="【売買損益】").font = Font(name="Meiryo", size=11, bold=True)
summary.cell(row=row, column=2, value="※総平均法で計算").font = Font(name="Meiryo", size=9, color="888888")
row = 4
for col_i, hdr in enumerate(["通貨", "売却額(円)", "売却原価(円)", "売買損益(円)"], 1):
    c = summary.cell(row=row, column=col_i, value=hdr)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border

trade_total = Decimal(0)
row = 5
for cur in currencies:
    # Total buy qty/amount (exchange + sec3 rewards count as acquisition)
    buy_qty = sum(d["bq"] for d in sec2[cur].values())
    buy_amt = sum(d["ba"] for d in sec2[cur].values())
    # Include sec3 acquisitions in cost pool
    for t in sec3.get(cur, []):
        buy_qty += t["bq"]
        buy_amt += t["ba"]
    sell_qty = sum(d["sq"] for d in sec2[cur].values())
    sell_amt = sum(d["sa"] for d in sec2[cur].values())
    if not sell_qty:
        continue  # 売却なし → 損益なし
    # 総平均法: average unit cost × sell qty = cost of goods sold
    if buy_qty > 0:
        avg_cost = buy_amt / buy_qty
        cost_of_sold = (avg_cost * sell_qty).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    else:
        cost_of_sold = Decimal(0)
    pnl = sell_amt - cost_of_sold
    trade_total += pnl
    summary.cell(row=row, column=1, value=cur).font = cell_font
    c = summary.cell(row=row, column=2, value=int(sell_amt))
    c.font = cell_font; c.number_format = num_fmt; c.border = thin_border
    c = summary.cell(row=row, column=3, value=int(cost_of_sold))
    c.font = cell_font; c.number_format = num_fmt; c.border = thin_border
    c = summary.cell(row=row, column=4, value=int(pnl))
    c.font = cell_font; c.number_format = num_fmt; c.border = thin_border
    for ci in range(1, 5):
        summary.cell(row=row, column=ci).border = thin_border
    row += 1

if trade_total == 0 and row == 5:
    summary.cell(row=row, column=1, value="（売却取引なし）").font = Font(name="Meiryo", size=10, color="888888")
    row += 1

# Trade subtotal
for ci in range(1, 5):
    summary.cell(row=row, column=ci).fill = sub_fill
    summary.cell(row=row, column=ci).border = thin_border
summary.cell(row=row, column=1, value="小計").font = Font(name="Meiryo", size=10, bold=True)
c = summary.cell(row=row, column=4, value=int(trade_total))
c.font = Font(name="Meiryo", size=10, bold=True); c.number_format = num_fmt; c.fill = sub_fill
row += 2

# --- Section B: 報酬（ステーキング・レンディング等）---
summary.cell(row=row, column=1, value="【報酬収入】").font = Font(name="Meiryo", size=11, bold=True)
summary.cell(row=row, column=2, value="※受取時に課税").font = Font(name="Meiryo", size=9, color="888888")
row += 1
for col_i, hdr in enumerate(["通貨", "内容", "数量", "金額(円)"], 1):
    c = summary.cell(row=row, column=col_i, value=hdr)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border

reward_total = Decimal(0)
row += 1
for cur in currencies:
    txs = sec3.get(cur, [])
    if not txs:
        continue
    # Group by description
    desc_agg = defaultdict(lambda: {"qty": Decimal(0), "amt": Decimal(0)})
    for t in txs:
        desc_agg[t["desc"]]["qty"] += t["bq"]
        desc_agg[t["desc"]]["amt"] += t["ba"]
    for desc, d in desc_agg.items():
        summary.cell(row=row, column=1, value=cur).font = cell_font
        summary.cell(row=row, column=2, value=desc).font = cell_font
        c = summary.cell(row=row, column=3, value=float(d["qty"]))
        c.font = cell_font; c.border = thin_border
        c = summary.cell(row=row, column=4, value=int(d["amt"]))
        c.font = cell_font; c.number_format = num_fmt; c.border = thin_border
        reward_total += d["amt"]
        for ci in range(1, 5):
            summary.cell(row=row, column=ci).border = thin_border
        row += 1

# Reward subtotal
for ci in range(1, 5):
    summary.cell(row=row, column=ci).fill = sub_fill
    summary.cell(row=row, column=ci).border = thin_border
summary.cell(row=row, column=1, value="小計").font = Font(name="Meiryo", size=10, bold=True)
c = summary.cell(row=row, column=4, value=int(reward_total))
c.font = Font(name="Meiryo", size=10, bold=True); c.number_format = num_fmt; c.fill = sub_fill
row += 2

# --- Grand total ---
for ci in range(1, 5):
    summary.cell(row=row, column=ci).fill = total_fill
    summary.cell(row=row, column=ci).border = thin_border
summary.cell(row=row, column=1, value="合計所得額").font = Font(name="Meiryo", size=12, bold=True)
grand = trade_total + reward_total
c = summary.cell(row=row, column=4, value=int(grand))
c.font = Font(name="Meiryo", size=12, bold=True)
c.number_format = num_fmt
c.fill = total_fill
row += 2

# Note
summary.cell(row=row, column=1, value="※注意").font = Font(name="Meiryo", size=9, bold=True)
row += 1
notes = [
    "・売買損益は総平均法で計算しています。年始残高がある場合は値が変わります。",
    "・年始残高がある場合は各計算書シートのセクション4を手動で修正してください。",
    "・報酬収入は受取時の時価で課税されます（売却時ではありません）。",
    f"・合計所得額が20万円以下のため、給与所得者は確定申告不要の可能性があります。",
]
for note in notes:
    summary.cell(row=row, column=1, value=note).font = Font(name="Meiryo", size=9, color="666666")
    summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

print(f"\n=== サマリーシート作成完了 ===")
print(f"  売買損益: {trade_total:,}円")
print(f"  報酬収入: {reward_total:,}円")
print(f"  合計所得: {grand:,}円")

wb.save(DST)
print(f"\n=== 保存完了: {DST} ===")
